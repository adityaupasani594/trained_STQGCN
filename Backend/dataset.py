"""
Vectorised ST-QGCN large dataset generator.
50 nodes, 8 weeks at 15-min resolution = 2,688 timesteps.
134,400 node-time records. Uses write_only xlsx for speed.
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta

np.random.seed(42); random.seed(42)

N_NODES=50; N_EDGES=90; T_STEPS=2688; TIME_RES_MIN=15
START_DT = datetime(2024,1,1,0,0)

# ── Node metadata ─────────────────────────────────────────────────────────────
node_ids = [f"N{i:03d}" for i in range(1,N_NODES+1)]
zones, types = [], []
for i in range(N_NODES):
    if i<17:   zones.append("CBD");         types.append("Signalized")
    elif i<35: zones.append("Residential"); types.append(random.choice(["Signalized","Roundabout"]))
    else:      zones.append("Highway");     types.append("Priority-based")

node_lat = np.random.uniform(19.05,19.35,N_NODES).round(6)
node_lon = np.random.uniform(72.80,73.15,N_NODES).round(6)
cap_map  = {"CBD":(800,1600),"Residential":(400,900),"Highway":(1200,2400)}
capacity = np.array([random.randint(*cap_map[z]) for z in zones], dtype=np.float32)

# ── Time arrays ───────────────────────────────────────────────────────────────
timestamps = [START_DT + timedelta(minutes=TIME_RES_MIN*t) for t in range(T_STEPS)]
ts_str     = [t.strftime("%Y-%m-%d %H:%M") for t in timestamps]
hours      = np.array([t.hour + t.minute/60 for t in timestamps], dtype=np.float32)
dows       = np.array([t.weekday() for t in timestamps], dtype=np.int8)

# ── Flow matrix [T, N] ────────────────────────────────────────────────────────
print("Generating flow matrix ...")
base_map = {"CBD":(400,900),"Residential":(150,450),"Highway":(600,1400)}
base_flows = np.array([random.uniform(*base_map[z]) for z in zones])

h_t   = hours[:,None]
am_pk = np.where(np.array(zones)=="Highway",8.0,np.where(np.array(zones)=="CBD",8.5,7.5))[None,:]
pm_pk = np.where(np.array(zones)=="Highway",17.0,np.where(np.array(zones)=="CBD",17.5,18.0))[None,:]
am_a  = np.where(np.array(zones)=="CBD",1.8,np.where(np.array(zones)=="Highway",1.6,1.4))[None,:]
pm_a  = np.where(np.array(zones)=="CBD",1.6,np.where(np.array(zones)=="Highway",1.5,1.3))[None,:]

rf = (1 + am_a*np.exp(-((h_t-am_pk)**2)/2.5) + pm_a*np.exp(-((h_t-pm_pk)**2)/3.0))
rf *= np.where(dows[:,None]>=5, 0.55, 1.0)
rf *= np.where((h_t<5)|(h_t>23), 0.25, 1.0)
weekly = (1 + 0.08*np.sin(2*np.pi*np.arange(T_STEPS)/(7*24*4)))[:,None]
noise  = np.random.normal(0, base_flows[None,:]*0.06, (T_STEPS,N_NODES)).astype(np.float32)
flow   = np.maximum(5.0, (base_flows[None,:]*rf*weekly+noise)).astype(np.float32)

# ── Incidents ─────────────────────────────────────────────────────────────────
print("Injecting incidents ...")
incident = np.zeros((T_STEPS,N_NODES),dtype=np.int8)
for _ in range(int(T_STEPS*N_NODES*0.002)):
    t0=random.randint(0,T_STEPS-12); ni=random.randint(0,N_NODES-1); dur=random.randint(4,12)
    incident[t0:t0+dur,ni]=1
    flow[t0:t0+dur,ni]*=np.float32(random.uniform(0.4,0.7))

# ── Propagation ───────────────────────────────────────────────────────────────
for _ in range(20):
    up,dn=random.randint(0,N_NODES-2),random.randint(0,N_NODES-1)
    w=random.uniform(0.08,0.20)
    delta=np.diff(flow[:,up],prepend=flow[0,up])*w*0.3
    flow[:,dn]=np.maximum(5.0,flow[:,dn]+delta)

# ── Derived signals ───────────────────────────────────────────────────────────
v_ff  = np.where(np.array(zones)=="Highway",80.0,50.0)[None,:]
ratio = np.clip(flow/capacity[None,:],0,1)
speed = np.maximum(5.0, v_ff*(1-ratio)+np.random.normal(0,2,(T_STEPS,N_NODES)).astype(np.float32))
occup = np.clip(ratio*100+np.random.normal(0,1,(T_STEPS,N_NODES)).astype(np.float32),0,99.9)
queue = np.maximum(0.0, np.where(ratio>0.7,(ratio-0.7)*capacity[None,:]*0.05,0)
                   +np.random.normal(0,2,(T_STEPS,N_NODES)).astype(np.float32))

# ── Build DataFrames ──────────────────────────────────────────────────────────
print("Building DataFrames ...")
t_idx = np.repeat(np.arange(T_STEPS),N_NODES)
n_idx = np.tile(np.arange(N_NODES),T_STEPS)

df_nodes = pd.DataFrame({
    "Timestamp":               pd.array([ts_str[t] for t in t_idx],dtype="string"),
    "Node_ID":                 pd.array([node_ids[n] for n in n_idx],dtype="string"),
    "Zone":                    pd.array([zones[n] for n in n_idx],dtype="string"),
    "Intersection_Type":       pd.array([types[n] for n in n_idx],dtype="string"),
    "Traffic_Flow_veh_per_hr": np.round(flow.ravel(),1),
    "Avg_Speed_kmh":           np.round(speed.ravel(),1),
    "Lane_Occupancy_pct":      np.round(occup.ravel(),2),
    "Queue_Length_m":          np.round(queue.ravel(),1),
})
print(f"  Node features  : {len(df_nodes):,} rows")

edges,seen=[],set()
while len(edges)<N_EDGES:
    s,t=random.randint(0,N_NODES-1),random.randint(0,N_NODES-1)
    if s!=t and (s,t) not in seen:
        seen.add((s,t)); edges.append((node_ids[s],node_ids[t]))

df_edge_static = pd.DataFrame([{
    "Edge_ID":f"{s}-{t}","Source_Node":s,"Target_Node":t,
    "Road_Capacity_veh_hr":random.randint(800,2400),
    "Segment_Length_m":random.randint(150,2500),
    "Road_Quality_IRI":round(random.uniform(1.2,9.0),2),
    "Speed_Limit_kmh":random.choice([30,40,50,60,80,100]),
    "Lane_Count":random.randint(1,4),
} for s,t in edges])

node_id_to_idx={nid:i for i,nid in enumerate(node_ids)}
e_t   = np.repeat(np.arange(T_STEPS),N_EDGES)
e_idx = np.tile(np.arange(N_EDGES),T_STEPS)
src_indices = np.array([node_id_to_idx[s] for s,_ in edges])
df_edge_ts = pd.DataFrame({
    "Timestamp": [ts_str[t] for t in e_t],
    "Edge_ID":   [f"{edges[e][0]}-{edges[e][1]}" for e in e_idx],
    "Incident_Flag": incident[:,src_indices].ravel().astype(int),
})
print(f"  Edge time series: {len(df_edge_ts):,} rows")

precip_arr=np.zeros(T_STEPS); p=0.0
for i in range(T_STEPS):
    p=max(0,p-0.05+np.random.normal(0,0.1)) if p>0 else (np.random.exponential(2.0) if random.random()<0.004 else 0)
    precip_arr[i]=max(0,p+np.random.normal(0,0.05))
df_weather = pd.DataFrame({
    "Timestamp":           ts_str,
    "Precipitation_mm_hr": np.round(precip_arr,2),
    "Visibility_m":        np.round(np.clip(1800-precip_arr*120+np.random.normal(0,40,T_STEPS),50,5000),1),
    "Temperature_C":       np.round(26+5*np.sin(2*np.pi*(hours-6)/24)+np.random.normal(0,0.8,T_STEPS),1),
})

HOLIDAYS={(1,26),(3,25),(4,14),(5,1),(8,15),(10,2),(12,25)}
df_temporal = pd.DataFrame({
    "Timestamp":     ts_str,
    "Hour_of_Day":   np.round(hours,2),
    "Hour_Sin":      np.round(np.sin(2*np.pi*hours/24),6),
    "Hour_Cos":      np.round(np.cos(2*np.pi*hours/24),6),
    "Day_of_Week":   dows.astype(int),
    "Is_Weekend":    (dows>=5).astype(int),
    "Holiday_Flag":  [1 if (t.month,t.day) in HOLIDAYS else 0 for t in timestamps],
    "Special_Event": [1 if (t.weekday()==5 and 18<=t.hour<=22) else 0 for t in timestamps],
})

df_topo = pd.DataFrame({
    "Node_ID":node_ids,"Zone":zones,"Latitude":node_lat,"Longitude":node_lon,
    "Intersection_Type":types,"Capacity_veh_hr":capacity.astype(int),
})

# ── Write Excel using pandas ExcelWriter (faster than openpyxl row-by-row) ───
print("Writing workbook (this takes ~30s for 134k rows) ...")
OUT = "STQGCN_Dataset_Large.xlsx"

with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    # README as a small df
    readme_df = pd.DataFrame([
        ("Nodes", N_NODES), ("Directed edges", N_EDGES),
        ("Timesteps", T_STEPS), ("Time resolution", f"{TIME_RES_MIN} min"),
        ("Duration", "8 weeks"), ("Node-time records", f"{N_NODES*T_STEPS:,}"),
        ("Edge-time records", f"{N_EDGES*T_STEPS:,}"),
        ("Zones", "CBD | Residential | Highway"), ("",""),
        ("TRAINING NOTES",""),
        ("seq_len recommended", "48–96  (12h–24h at 15-min resolution)"),
        ("Available windows (seq=96)", f"~{T_STEPS-96-1:,}"),
        ("Train windows ~80%", f"~{int((T_STEPS-96-1)*0.8):,}"),
        ("Val windows ~20%",   f"~{int((T_STEPS-96-1)*0.2):,}"),
        ("Run command","python train_stqgcn.py --data STQGCN_Dataset_Large.xlsx --seq-len 96 --epochs 100"),
    ], columns=["Parameter","Value"])
    readme_df.to_excel(writer, sheet_name="README", index=False)

    df_nodes.to_excel(writer,       sheet_name="Node_Features",    index=False)
    df_edge_static.to_excel(writer, sheet_name="Edge_Static",      index=False)
    df_edge_ts.to_excel(writer,     sheet_name="Edge_Time_Series", index=False)
    df_weather.to_excel(writer,     sheet_name="Weather_Features", index=False)
    df_temporal.to_excel(writer,    sheet_name="Temporal_Features",index=False)
    df_topo.to_excel(writer,        sheet_name="Node_Topology",    index=False)

    # Style headers post-write
    wb = writer.book
    SHEET_COLORS = {
        "README":           "1F4E79",
        "Node_Features":    "1F4E79",
        "Edge_Static":      "375623",
        "Edge_Time_Series": "375623",
        "Weather_Features": "7B2D8B",
        "Temporal_Features":"833C00",
        "Node_Topology":    "2E75B6",
    }
    for sname, color in SHEET_COLORS.items():
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        fill  = PatternFill("solid", fgColor=color)
        font  = Font(bold=True, color="FFFFFF", size=10)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill=fill; cell.font=font; cell.alignment=align
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

print(f"\nSaved → {OUT}")
print(f"Node-time rows  : {N_NODES*T_STEPS:,}")
print(f"Edge-time rows  : {N_EDGES*T_STEPS:,}")
print(f"Train windows   : ~{int((T_STEPS-96-1)*0.8):,}  (seq_len=96)")
print(f"\nRun:")
print(f"  python train_stqgcn.py --data STQGCN_Dataset_Large.xlsx --seq-len 96 --epochs 100")